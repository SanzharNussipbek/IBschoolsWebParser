import requests
from lxml import etree
import lxml.html
import openpyxl
import time
import pandas as pd
import string
import json


# function to abbreviate programme names
def abbreviate(word):
    if word == 'Diploma Programme':
        return 'DP'
    elif word == 'Primary Years Programme':
        return 'PYP'
    elif word == 'Middle Years Programme':
        return 'MYP'
    elif word == 'Career-related Programme':
        return 'CP'

# function to turn list of words into one string separated by a comma or a space
def to_string(values, comma):
    result = ''
    for i in range(len(values)):
        result += string.capwords(values[i])
        if i != len(values)-1:
            result += ', ' if comma else ' '
    return result

# check if given string has numbers
def hasNumbers(inputString):
    return any(char.isdigit() for char in inputString)


# function to return the DP Coordinator values from the list of several coordinators
def get_DP_coordinator(data):
    if len(data) == 1:
        return data[0]
    else:
        titles = []
        for item in data:
            titles.append(item['title'])
        
        if 'Diploma Programme' in titles:
            index = titles.index('Diploma Programme')
            data[index]['title'] = 'DP Coordinator'
        elif 'Middle Years Programme' in titles:
            index = titles.index('Middle Years Programme')
            data[index]['title'] = 'MYP Coordinator'
        elif 'Primary Years Programme' in titles:
            index = titles.index('Primary Years Programme')
            data[index]['title'] = 'PYP Coordinator'

        return data[index]


# function to return the address in a proper format
def get_proper_address(values):
    address = ''
    for i in range(len(values)):
        values[i] = values[i].strip().replace('\r', '').replace('\n', '')
        
        if values[i].isupper() and not hasNumbers(values[i]):   # check if the word is not ZIP Code 
                values[i] = values[i].capitalize()
        
        address += values[i]
        
        if i != len(values)-1:
            address += ', '
    
    return address


# function to get information of coordinators, there can be up to 4 coordinators : PYP, MYP, DP, CP
def get_coordinator(url):
    api = requests.get(url)
    tree = lxml.html.document_fromstring(api.text)
    
    path_names = '//*[@id="mainContent"]/div[2]/div/div[1]/aside/p/text()'                  # names of the coordinators
    path_phones = '//*[@id="mainContent"]/div[2]/div/div[1]/aside/ul/li[2]/span[2]/text()'  # phones of the coordinators
    path_title = '//*[@id="mainContent"]/div[2]/div/div[1]/aside/h3[{}]/text()'             # titles of the coordinators
    path_address = '//*[@id="mainContent"]/div[2]/div/div[1]/aside/address[{}]/p/text()'    # addresses of the schools
    
    names = tree.xpath(path_names)
    phones = tree.xpath(path_phones)
    titles = []
    addresses =[]

    for i in range(1,5):
        titles.append(tree.xpath(path_title.format(i)))
        addresses.append(tree.xpath(path_address.format(i)))
    
    coordinators = []

    for i in range(len(names)):
        title = titles[i][0]  if len(titles[0]) != 0 else 'UNKNOWN'
        name_vals = names[i].split(' ')
        address = get_proper_address(addresses[i])
        phone = 'UNKNOWN' if len(phones) != len(names) else phones[i].replace('-', '').replace(' ', '').replace('(', '').replace(')', '').replace('=', '').replace('+', '')
        
        coordinators.append({
            'salute': name_vals[0].capitalize(),
            'name' : to_string(name_vals[1:-1], False),
            'surname' : name_vals[-1].capitalize(),
            'phone': phone,
            'title' : title,
            'address' : address
        })

    DP_coordinator = get_DP_coordinator(coordinators)
    return DP_coordinator


# function to get the webstite url of the school
def get_website(url):
    path = '//*[@id="mainContent"]/div[2]/div/div[1]/div/div[1]/div[7]/div[2]/a/@href'
    path2 = '//*[@id="mainContent"]/div[2]/div/div[1]/div/div[1]/div[6]/div[2]/a/@href'
    
    api = requests.get(url)
    tree = lxml.html.document_fromstring(api.text)
    
    website = tree.xpath(path)
    website = website[0] if len(website) != 0 else tree.xpath(path2)[0]

    return website

# function to get the country name
def get_country_name(url):
    path = '//*[@id="mainContent"]/div[2]/div/div[1]/div/div[1]/div[4]/div[2]/text()'
    path2 = '//*[@id="mainContent"]/div[2]/div/div[1]/div/div[1]/div[3]/div[2]/text()'

    api = requests.get(url)
    tree = lxml.html.document_fromstring(api.text)
    
    country_name = tree.xpath(path)
    country_name = country_name[0] if len(country_name) != 0 else 'UNKNOWN'

    if country_name[:2] == 'IB':
        print('hello')
        country_name = tree.xpath(path2)
        country_name = country_name[0] if len(country_name) != 0 else 'UNKNOWN'
    
    return country_name

# function to parse the given url
def parse(url):
    api = requests.get(url)
    tree = lxml.html.document_fromstring(api.text)
    
    names = tree.xpath('//*[@id="mainContent"]/div[3]/div/div/table/tbody/tr/td[1]/a/text()')  # names of the schools in the page
    curriculum_path = '//*[@id="mainContent"]/div[3]/div/div/table/tbody/tr[{}]/td/span/@data-tooltip' # curriculums of the schools
    links = tree.xpath('//*[@id="mainContent"]/div[3]/div/div/table/tbody/tr/td[1]/a/@href') # links to the schools' pages
    
    data = []
    HOST = 'https://www.ibo.org' # HOST link
    
    for i in range(len(names)):
        curriculum = tree.xpath(curriculum_path.format(i+1))    # get curriculum of the school
        link = HOST + links[i]                                  # get the link of the school's personal page by appending the HOST link
        contact = get_coordinator(link)                         # get DP coordinator of the school from the school's personal page
        data.append({                                           # append the data list with school's information
            'country' : get_country_name(link),
            'name' : string.capwords(names[i]),
            'curriculum' : 'IB',
            'link' : link,
            'address' : contact['address'],
            'website' : get_website(link),
            'salutation' : contact['salute'],
            'contact_name' : contact['name'],
            'contact_surname' : contact['surname'],
            'contact_phone' : contact['phone'],
            'contact_title' : 'DP Coordinator' if contact['title'] == 'UNKNOWN' else contact['title']
        })

    return data


# function to save all the parsed data to excel file
def save(data):
    # open new excel file 
    wb = openpyxl.Workbook()
    wb.remove(wb['Sheet'])
    wb.create_sheet(title = 'Page 0')
    sheet = wb['Page 0']
    rownum = 2
    
    # headers for the excel table
    headers = [
                'COUNTRY', 
                'CITY', 
                'NAME', 
                'ADDRESS', 
                'POSTAL CODE', 
                'SALUTATION', 
                'CONTACT NAME', 
                'CONTACT SURNAME', 
                'CONTACT TITLE', 
                'CONTACT EMAIL', 
                'SCHOOL EMAIL', 
                'COUNTRY CODE 1', 
                'CONTACT PHONE', 
                'COUNTRY CODE 2', 
                'SCHOOL PHONE', 
                'WEBSITE', 
                'CURRICULUM', 
                'LINK']

    # put the headers into the table
    for i in range(1, len(headers)+1):
        cell = sheet.cell(row = 1, column = i)
        cell.value = headers[i-1]
    
    print('\nInputting data to file.....')

    # iterate through the data and input to the columns
    for item in data:
        for value in item:
            cell1 = sheet.cell(row = rownum, column = 1)
            cell1.value = value['country']

            cell3 = sheet.cell(row = rownum, column = 3)
            cell3.value = value['name']

            cell4 = sheet.cell(row = rownum, column = 4)
            cell4.value = value['address']

            cell6 = sheet.cell(row = rownum, column = 6)
            cell6.value = value['salutation']

            cell7 = sheet.cell(row = rownum, column = 7)
            cell7.value = value['contact_name']

            cell8 = sheet.cell(row = rownum, column = 8)
            cell8.value = value['contact_surname']

            cell9 = sheet.cell(row = rownum, column = 9)
            cell9.value = value['contact_title']

            cell13 = sheet.cell(row = rownum, column = 13)
            cell13.value = value['contact_phone']

            cell16 = sheet.cell(row = rownum, column = 16)
            cell16.value = value['website']

            cell17 = sheet.cell(row = rownum, column = 17)
            cell17.value = value['curriculum']

            cell18 = sheet.cell(row = rownum, column = 18)
            cell18.value = value['link']
            
            rownum+=1
    
    # save the data to the file
    wb.save('IB_Schools.xlsx')
    print('Done!')


# function to turn list of the countries and their information to a JSON file
def init_json():
    # [country_code, number of pages, country name]
    country_list = [
                    # Americas
                    ['AG', 1, 'Antigua and Barbuda'],
                    ['AR', 2, 'Argentina'],
                    ['BS', 1, 'Bahamas'],
                    ['BB', 1, 'Barbados'],
                    ['BM', 1, 'Bermuda'],
                    ['BO', 1, 'Bolivia'],
                    ['BR', 2, 'Brazil'],
                    ['CA', 9, 'Canada'],
                    ['KY', 1, 'Cayman Islands'], 
                    ['CL', 1, 'Chile'], 
                    ['CO', 2, 'Colombia'],
                    ['CR', 2, 'Costa Rica'], 
                    ['CW', 1, 'Curacao'], 
                    ['DO', 1, 'Dominican Republic'],
                    ['EC', 2, 'Ecuador'],
                    ['SV', 1, 'El Salvador'],
                    ['GT', 1, 'Guatemala'],
                    ['HN', 1, 'Honduras'],
                    ['JM', 1, 'Jamaica'],
                    ['MX', 3, 'Mexico'],
                    ['NI', 1, 'Nicaragua'],
                    ['PA', 1, 'Panama'],
                    ['PY', 1, 'Paraguay'],
                    ['PE', 2, 'Peru'],
                    ['PR', 1, 'Puerto Rico'],
                    ['SX', 1, 'Sint Maarten (dutch)'],
                    ['UY', 1, 'Uruguay'],
                    ['US', 49, 'USA'],
                    ['VE', 1, 'Venezuela'],
                    ['VG', 1, 'Virgin Islands, British'],
                    ['VI', 1, 'Virgin Islands, U.S.'],
                    # Africa, Europe and Middle East
                    ['AZ', 1, 'Azerbaijan'],
                    ['AL', 1, 'Albania'],
                    ['AD', 1, 'Andorra'],
                    ['AO', 1, 'Angola'],
                    ['AM', 1, 'Armenia'],
                    ['AT', 1, 'Austria'],
                    ['BH', 1, 'Bahrain'],
                    ['BE', 1, 'Belgia'],
                    ['BA', 1, 'Bosnia and Herzegovina'],
                    ['BW', 1, 'Botswana'],
                    ['BG', 1, 'Bulgaria'],
                    ['BF', 1, 'Burkina Faso'],
                    ['CM', 1, 'Cameroon'],
                    ['CD', 1, 'Congo'],
                    ['CI', 1, "Cote D'ivoire"],
                    ['HR', 1, "Croatia"],
                    ['CU', 1, "Cuba"],
                    ['CY', 1, "Cyprus"],
                    ['CZ', 1, "Czech Republic"],
                    ['DK', 1, "Denmark"],
                    ['ER', 1, "Eritrea"],
                    ['EE', 1, "Estonia"],
                    ['ET', 1, "Ethiopia"],
                    ['FI', 1, "Finlandia"],
                    ['FR', 1, "France"],
                    ['GA', 1, 'Gabon'],
                    ['GE', 1, "Georgia"],
                    ['DE', 4, "Germany"],
                    ['GH', 1, "Ghana"],
                    ['GR', 1, "Greece"],
                    ['GG', 1, "Guernsey"],
                    ['HU', 1, "Hungary"],
                    ['IS', 1, "Iceland"],
                    ['IR', 1, "Iran"],
                    ['IQ', 1, "Iraq"],
                    ['IE', 1, "Ireland"],
                    ['IM', 1, "Isle of Man"],
                    ['IL', 1, "Israel"],
                    ['IT', 2, "Italy"],
                    ['JE', 1, "Jersey"],
                    ['JO', 1, 'Jordan'],
                    ['KZ', 1, 'Kazakhstan'],
                    ['KE', 1, 'Kenia'],
                    ['KW', 1, 'Kuwait'],
                    ['KG', 1, "Kyrgyzstan"],
                    ['LV', 1, "Latvia"],
                    ['LB', 1, "Lebanon"],
                    ['LS', 1, "Lesotho"],
                    ['LT', 1, "Lithuania"],
                    ['LU', 1, "Luxemburg"],
                    ['MG', 1, "Madagaskar"],
                    ['MW', 1, "Malawi"],
                    ['ML', 1, "Mali"],
                    ['MT', 1, "Malta"],
                    ['MU', 1, 'Mauritius'],
                    ['MC', 1, "Monaco"],
                    ['ME', 1, "Montenegro"],
                    ['MZ', 1, "Mazambique"],
                    ['NA', 1, 'Namibia'],
                    ['NG', 1, 'Nigeria'],
                    ['NL', 1, "Netherlands"],
                    ['NO', 2, "Norway"],
                    ['OM', 1, "Oman"],
                    ['PS', 1, "Palestine"],
                    ['PL', 3, "Poland"],
                    ['PT', 1, "Portugal"],
                    ['MK', 1, "North Macedonia"],
                    ['QA', 1, 'Qatar'],
                    ['RU', 2, 'Russia'],
                    ['RO', 1, "Romania"],
                    ['RW', 1, "Rwanda"],
                    ['SA', 1, 'Saudi Arabia'],
                    ['ZA', 1, 'South Africa'],
                    ['SN', 1, "Senegal"],
                    ['RS', 1, "Serbia"],
                    ['SK', 1, "Slovakia"],
                    ['SI', 1, "Slovenia"],
                    ['ES', 4, "Spain"],
                    ['SD', 1, "Sudan"],
                    ['SZ', 1, "Swaziland"],
                    ['SE', 2, "Sweden"],
                    ['CH', 3, "Switzerland"],
                    ['TZ', 1, "Tanzania"],
                    ['TG', 1, "Togo"],
                    ['TN', 1, "Tunisia"],
                    ['TR', 3, 'Turkey'],
                    ['UG', 1, 'Uganda'],
                    ['UA', 1, 'Ukraine'],
                    ['AE', 3, 'UAE'],
                    ['UZ', 1, 'Uzbekistan'],
                    ['GB', 5, "United Kingdom"],
                    ['ZM', 1, 'Zambia'],
                    ['ZW', 1, 'Zimbabwe'],
                    # Asia Pacific
                    ['AU', 4, 'Australia'],
                    ['BD', 1, 'Bangladesh'],
                    ['BN', 1, 'Brunei Darussalam'],
                    ['KH', 1, 'Cambodia'],
                    ['CN', 6, 'China'],
                    ['FJ', 1, 'Fiji'],
                    ['GU', 1, 'Guam'],
                    ['HK', 2, 'Hong Kong'],
                    ['IN', 8, 'India'],
                    ['ID', 3, 'Indonesia'],
                    ['JP', 3, 'Japan'],
                    ['KR', 1, 'Korea'],
                    ['LA', 1, 'Lao'],
                    ['MO', 1, 'Macao'],
                    ['MY', 1, 'Malaysia'],
                    ['MN', 3, 'Mongolia'],
                    ['MM', 1, 'Myanmar'],
                    ['NP', 1, 'Nepal'],
                    ['NZ', 1, 'New Zealand'],
                    ['PK', 1, 'Pakistan'],
                    ['PG', 1, 'Papua New Guinea'],
                    ['PH', 1, 'Philippines'],
                    ['SG', 2, 'Singapore'],
                    ['LK', 1, 'Sri Lanka'],
                    ['TW', 1, 'Taiwan'],
                    ['TH', 2, 'Thailand'],
                    ['VN', 1, 'Vietnam']
                    ]


    # save countries by regions

    americas = []
    for item in country_list[:31]:
        country_dict = {}
        country_dict["country_code"] = item[0]
        country_dict["num_of_pages"] = item[1]
        country_dict["country_name"] = item[2]
        americas.append(country_dict)

    asia = []
    for item in country_list[31:124]:
        country_dict = {}
        country_dict["country_code"] = item[0]
        country_dict["num_of_pages"] = item[1]
        country_dict["country_name"] = item[2]
        asia.append(country_dict)

    africa = []
    for item in country_list[124:]:
        country_dict = {}
        country_dict["country_code"] = item[0]
        country_dict["num_of_pages"] = item[1]
        country_dict["country_name"] = item[2]
        africa.append(country_dict)

    # initialize dictionary of countries and regions
    countries = {'Americas' : americas, 'Asia Pacific' : asia, 'Africa, Europe and Middle East' : africa}
    
    # Serializing json  
    json_object = json.dumps(countries, indent = 4) 
    
    # Writing to sample.json 
    with open("countries.json", "w") as outfile: 
        outfile.write(json_object) 


# main function
def main():
    init_json() # create json file

    # dictionary of the urls for three different regions
    # the urls have two places to be formatted: country code and page number
    urls = {
            'Americas' : 'https://www.ibo.org/programmes/find-an-ib-school/?SearchFields.Region=iba&SearchFields.Country={}&SearchFields.Keywords=&SearchFields.Language=English&SearchFields.BoardingFacilities=&SearchFields.SchoolGender=&SearchFields.ProgrammeDP=true&page={}',
            'Asia Pacific' : 'https://www.ibo.org/programmes/find-an-ib-school/?SearchFields.Region=ibap&SearchFields.Country={}&SearchFields.Keywords=&SearchFields.Language=English&SearchFields.BoardingFacilities=&SearchFields.SchoolGender=&SearchFields.ProgrammeDP=true&page={}',
            'Africa, Europe and Middle East' : 'https://www.ibo.org/programmes/find-an-ib-school/?SearchFields.Region=ibaem&SearchFields.Country={}&SearchFields.Keywords=&SearchFields.Language=English&SearchFields.BoardingFacilities=&SearchFields.SchoolGender=&SearchFields.ProgrammeDP=true&page={}'
            }

    # load the data from JSON file
    with open('countries.json') as json_file:
        country_list = json.load(json_file)
    
    # inititalize empty data list
    data = []

    # iterate through data and parse the url
    # for region in country_list:
    #     print('\nParsing {}...'.format(region))

    #     for country in country_list[region]:
    #         print('\n\tParsing {}...'.format(country["country_name"]))

    #         for i in range(1, country["num_of_pages"]+1):
    #             print('\t\tParsing Page {} out of {}.....'.format(i, country["num_of_pages"]))
    #             data.append(parse(urls[region].format(country["country_code"], i)))


    data.append(parse(urls['Americas'].format('CA', 1)))

    
    # save the data to excel file
    save(data)


if __name__ == "__main__":
    main()